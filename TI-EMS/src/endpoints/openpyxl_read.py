import subprocess
from openpyxl import load_workbook

# Define file paths
input_file_path = r'D:\thowffic\TI-EMS\src\endpoints\attachments\example_boss.xlsx'
output_file_path = r'D:\thowffic\TI-EMS\src\endpoints\attachments\example_boss_file.xlsx'

# Example command (replace with your actual paths)
excel_command = f'"excel.exe" /r "{input_file_path}" /t "{output_file_path}"'

# Execute the command using subprocess
subprocess.run(excel_command, shell=True, check=True)

# Load the workbook
workbook = load_workbook(output_file_path, data_only=True)

# Select the worksheet
worksheet = workbook['Sheet']

# Access the cell
cell = worksheet['A3']

# Read the value of the cell (not the formula)
cell_value = cell.value

# Print the cell value
print(cell_value)



import openpyxl
import subprocess

# Define the path to your Excel file
excel_file_path = 'C:/path/to/your/excel_file.xlsx'

# Load the workbook
wb = openpyxl.load_workbook(excel_file_path)

# Perform any operations you need on the workbook
# Example: accessing a sheet and modifying a cell value
sheet = wb.active
sheet['A1'] = 'Updated Value'

# Save the workbook
wb.save(excel_file_path)

# Define the command you want to run in the terminal
command = "dir"

# Run the command
subprocess.run(command, shell=True)

# Close the workbook (Not necessary since we're not keeping it open, but good practice)
wb.close()

print("Excel file updated, command executed, and workbook closed.")
