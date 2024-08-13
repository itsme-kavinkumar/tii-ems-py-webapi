from openpyxl import load_workbook
import json
import os

# Example command (replace with your actual paths)
excel_command = r'"excel.exe" /r "D:\thowffic\TI-EMS\src\endpoints\attachments\example_boss.xlsx" /t "D:\thowffic\TI-EMS\src\endpoints\attachments\example_boss_file.xlsx"'

# Execute the command
os.system(excel_command)

# Load the workbook
# workbook = load_workbook('D:\\thowffic\\TI-EMS\\src\\endpoints\\attachments\\availability_report.xlsx', data_only=True)
workbook = load_workbook('D:\\thowffic\\TI-EMS\\src\\endpoints\\attachments\\example_boss_file.xlsx', data_only=True)
# workbook = load_workbook('D:\\thowffic\\TI-EMS\\example_boss.xlsx', data_only=True)

# Select the worksheet
worksheet = workbook['Sheet']

# Access the cell
cell = worksheet['A3']

# Read the value of the cell (not the formula)
cell_value = cell.value

print(cell_value)

start_row = 39
end_row = 62 - 1
start_col = 1
end_col = 6



pre_define_words = {1: 'ac', 2: 'desc', 5: 'dr', 6: 'cr'}  

account_lst = []

for row in range(start_row, end_row + 1):
    inner_dict = {}
    for col in range(start_col, end_col + 1):
        if col not in [3, 4]:
            cell = worksheet.cell(row, col)
            cell_value = cell.value
            print(cell_value)
            if col != 2:
                result = cell_value if cell_value else 0
                inner_dict[pre_define_words[col]] = int(result) if isinstance(result, float) else result
            else:
                cell_value = cell_value if cell_value else 0
                inner_dict[pre_define_words[col]] = f"{cell_value} - 123" if col == 2 else cell_value

    if inner_dict.get('dr', 0) != 0 or inner_dict.get('cr', 0) != 0:
        account_lst.append(inner_dict)

json_string = json.dumps(account_lst)
oracle_data = '{"ems":'+json_string+'}'
print(oracle_data)