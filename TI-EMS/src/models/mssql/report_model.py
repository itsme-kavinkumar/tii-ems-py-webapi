from sqlalchemy import text
from src.endpoints.response_json import _getReturnResponseJson,_getSuccessResponseJson,_getErrorResponseJson,get_exception_response
import json
from src.models.image import parse_date
from datetime import date
import os
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pathlib import Path
import calendar
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from log_file import createFolder
import re
from openpyxl.drawing.image import Image

static_dir = Path(__file__).parent 
base_path = Path(__file__).parent / "attachment"


def month_report(cnx,employee_id,machine_id,month_year,report_for,report_type,request):
    try:
        
        groupby = ""
        where = ""
        result_query = ''
        department_id = ''
        shed_id =''
        machinetype_id = ''
        rslt = ''
        
        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
        
        machine_id = id(machine_id)
        
        if machine_id == "":
            pass
        else:
    
            where += f" and mm.machine_id IN ({machine_id})"

        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()

            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]

        if department_id !='' and department_id !=0:
            where += f" and md.department_id ={department_id}"

        if shed_id !='' and shed_id != 0:
            where += f" and ms.shed_id ={shed_id}"

        if machinetype_id !='' and machinetype_id!= 0:
            where += f" and mmt.machinetype_id ={machinetype_id}"

        if report_for == '6to6':
            month, year = month_year.split('-')
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month}{year}'"""
            result_query = cnx.execute(query).mappings().all()

            if len(result_query)==0:
                return _getErrorResponseJson("power table not available...")
            tbl_name = f"ems_v1_completed.dbo.power_{month}{year} cp" 

        else:
            month, year = month_year.split('-')
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month}{year}_12'"""
            result_query = cnx.execute(query).mappings().all()

            if len(result_query)==0:
                return _getErrorResponseJson("power table not available...") 
               
            tbl_name = f"ems_v1_completed.dbo.power_{month}{year}_12 cp"
            if report_type not in ['date']:
                 return _getErrorResponseJson("Invalid report type")

        if report_type not in ['date', 'shift']:
                 return _getErrorResponseJson("Invalid report type")

        if report_type == 'date':
            day = "CONCAT('d', DAY(cp.mill_date))"
            
        elif report_type == 'shift':
            day = "CONCAT(CONCAT('ds', cp.mill_shift), '_', DAY(cp.mill_date))"
            groupby = ",cp.mill_shift"

        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                {day} AS day,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh
            FROM
                {tbl_name}
                INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                left join ems_v1.dbo.master_machine_factor mmf on mmf.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id                   
                LEFT JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id                   
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id                   
            WHERE
                1=1 {where} and FORMAT(cp.mill_date, 'MM-yyyy') = '{month_year}' 
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                DAY(cp.mill_date)
                {groupby}               
        ''')
        rslt = cnx.execute(query).mappings().all()
        print(query)
        if rslt !='':
            output = {}

            if report_type == 'date':
                output_keys = [f'd{day}' for day in range(1, 32)]
            elif report_type == 'shift':
                output_keys = [f'ds{shift}_{day}' for day in range(1, 32) for shift in range(1, 4)]
            
            for row in rslt:
                machine_code = row.machine_code
                machine_name = row.machine_name
                day = row.day
                kwh = row.kwh
            
                if machine_code not in output:
                    output[machine_code] = {
                        'machine_code': machine_code,
                        'machine_name': machine_name
                    }
                    for key in output_keys:
                        output[machine_code][key] = 0
            
                output[machine_code][day] = kwh
                result=list(output.values())
            if rslt == []:
                result = []
               
            # machine = {"result" : data["machine_name"],
            #            "result1" : data[]}
        return result
    except Exception as e:
        return get_exception_response(e)

def daily_report(cnx,date,report_for,request):
    try:
        res = ''
        ress = ''
        res_q4 = ''
        formula_d = 0
        formula_m = 0
        formula_y = 0
        func_name = ''
        formula = ''
        roundoff_value = 0
        type = ''
        dates=parse_date(date)
        if report_for == "":
            return _getErrorResponseJson("report_for is required")

        from_date_str = dates.strftime("%d-%m-%Y")
        datetime_obj = datetime.strptime(from_date_str, "%d-%m-%Y")
        f_date = datetime_obj.strftime("%d-%m-%Y") 
        formatted_date = datetime_obj.strftime("%d-%b-%Y") 
        month = f_date[3:5]
        given_year = formatted_date[7:]
        print("month",month)
        if int(month) >=4:
            year = formatted_date[7:]
            next_year = int(given_year) +1
        if int(month) <4:
            year = int(given_year) -1
            next_year =formatted_date[7:]
        print("next_year",next_year)

        query = f'''SELECT * FROM [ems_v1].[dbo].[master_energy_calculations] ORDER BY s_no '''
        result = cnx.execute(query).mappings().all()

        para = ''
        mill_month={1:"01",2:"02",3:"03",4:"04",5:"05",6:"06",7:"07",8:"08",9:"09",10:"10",11:"11",12:"12"}
        completed_db="[ems_v1_completed].[dbo]."    
        table_name = ""
        where = ""
        if report_for == '6to6':
            for rows in result:
                para = rows['parameter']
                    
            if para == 'kw':
                para = "case when mmf.kw = '*' then p.t_watts * mmf.kw_value when  mmf.kw = '/' then p.t_watts / mmf.kw_value else p.t_watts end "

            if para == 'kWh':
                para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
            
            month_year=f"""{mill_month[dates.month]}{str(dates.year)}"""
            table_name=f"  {completed_db}[power_{month_year}] as p"
            tblname = f'power_{month_year}'
            query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{tblname}'"""
            print(query1)
        else:
            para = "case when mmf.kWh = '*' then p.kWh * mmf.kWh_value when  mmf.kWh = '/' then p.kWh / mmf.kWh_value else p.kWh end "
            
            month_year=f"""{mill_month[dates.month]}{str(dates.year)}"""
            table_name=f"  {completed_db}[power_{month_year}_12] as p"
            tblname = f'power_{month_year}_12'
            query1 = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{tblname}'"""
            print(query1)
        query1 = cnx.execute(query1).mappings().all()
        where += f'''p.mill_date = '{dates}' '''   
       
        if len(query1)>0:     
            
            query2= f''' 
            SELECT 
                p.machine_id,
                sum({para}) as kWh,
                SUM(case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end ) AS reverse_kwh,
                min(mm.machine_name) as machine_name
            from {table_name} 
                left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
            where 
                {where} 
            group by 
                p.machine_id 
            order by 
                p.machine_id '''
            createFolder("Log/","query for day "+str(query2))
            res = cnx.execute(query2).mappings().all()

            dict={}
            dict_r_d={}
            
            for row in res:
                dict[row['machine_id']] = row['kWh']
                dict_r_d[row['machine_id']] = row['reverse_kwh']
              
            query3= f''' 
            SELECT 
                p.machine_id,
                sum({para}) as total ,
                SUM(case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end ) AS reverse_kwh,
                min(mm.machine_name) as machine_name
            from 
                {table_name}  
                left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id
                left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id
            group by 
                p.machine_id 
            order by 
                p.machine_id '''
            ress = cnx.execute(query3).mappings().all()
            createFolder("Log/","query for month "+str(query3))
            dict1={}
            dict_r_m = {}
            
            for row in ress:
                dict1[row['machine_id']] = row['total']
                dict_r_m[row['machine_id']] = row['reverse_kwh']   
        else:
            pass
        
        table_names = []
        # for month in range(1, 13):
        #     month_year = f"{mill_month[month]}{str(dates.year)}"
        #     if report_for == '6to6':
        #         tblname = f'power_{month_year}'
        #     else:
        #         tblname = f'power_{month_year}_12'
        #     query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = '{tblname}'"""
        #     result_query = cnx.execute(query).mappings().all()

        #     if len(result_query) > 0:
        #         if report_for == '6to6':
        #             tblname = f'power_{month_year}'
        #         else:
        #             tblname = f'power_{month_year}_12'
        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"

            if report_for == "12to12":
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all()
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all()
            if len(result_query) > 0:
                table_names.append(f"ems_v1_completed.dbo.power_{month_year}")
        
        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"

            if report_for == "12to12":
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all() 
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all() 
            if len(result_query) > 0:        
                table_names.append(f"ems_v1_completed.dbo.power_{month_year}")

        if len(table_names)==0:
            return _getErrorResponseJson("table not available")
        if report_for =="12to12":
            type = f"_12 p"
        else:
            type = f" p"
        
        union_query = " UNION ALL ".join([f"SELECT p.machine_id, {para} as data, case when mmf.kWh = '*' then p.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then p.reverse_kwh / mmf.kWh_value else p.reverse_kwh end  AS reverse_kwh, mm.machine_name FROM {table_name}{type}  left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=p.machine_id left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id" for table_name in table_names])

        query4 = f"""
                SELECT 
                    pp.machine_id, 
                    SUM(pp.data) as total_kwh ,
                    sum(pp.reverse_kwh) as reverse_kwh ,
                    min(pp.machine_name) as machine_name
                FROM 
                    ({union_query}) AS pp 
                    
                GROUP BY 
                    pp.machine_id 
                ORDER BY 
                    pp.machine_id"""
        createFolder("Log/","query for year "+str(query4))
        res_q4 = cnx.execute(query4).mappings().all()
        
        dict2={}
        dict_r_y = {}
        
        for row in res_q4:
            dict2[row['machine_id']] = row['total_kwh']
            dict_r_y[row['machine_id']] = row['reverse_kwh']
                     
        rows_to_write = []

        for rows in result:
            func_name = rows['function_name']
            formula = rows['formula2']
            roundoff_value = rows['roundoff_value']
            
            if len(res) == 0:
                # day_resultss = 0
                formula_d = 0
            else:
                if func_name == 'Power Import':
                    # day_resultss = dict[14]
                    formula_d = dict[14]

                elif func_name == 'Power Export':
                    # day_resultss = dict_r_d[14]
                    formula_d = dict_r_d[14]
                else:
                    # day_resultss = eval(formula, {"dict": dict})
                    
                    numbers = re.findall(r'\[(\d+)\]', formula)
                    valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict]
                    numeric_formula = formula
                    for machine_id in valid_ids:
                        numeric_value = dict.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                        numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                    formula_d = re.sub(r'dict', '', numeric_formula)

            if len(ress) == 0:
                month_resultsss = 0
                formula_m = 0
            else:
                if func_name == 'Power Import':
                    # month_resultsss = dict1[14]
                    formula_m = dict1[14]
                elif func_name == 'Power Export':
                    # month_resultsss = dict_r_m[14]
                    formula_m = dict_r_m[14]
                else:
                    # month_resultsss = eval(formula, {"dict": dict1})
                    numbers = re.findall(r'\[(\d+)\]', formula)

                    valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict1]
                    numeric_formula = formula
                    for machine_id in valid_ids:
                        numeric_value = dict1.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                        numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                    formula_m = re.sub(r'dict', '', numeric_formula)

            if len(res_q4) == 0:
                year_resultsss = 0
                formula_y = 0
            else:
                if func_name == 'Power Import':
                    # year_resultsss = dict2[14]
                    formula_y = dict2[14]
                    print("formula_y",formula_y)
                elif func_name == 'Power Export':
                    # year_resultsss = dict_r_y[14]
                    formula_y = dict_r_y[14]

                else:
                    # year_resultsss = eval(formula, {"dict": dict2})
                    
                    numbers = re.findall(r'\[(\d+)\]', formula)

                    valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict2]
                    numeric_formula = formula
                    for machine_id in valid_ids:
                       
                        numeric_value = dict2.get(machine_id, 0)  # Get the value from dict2 or use 0 if not found
                        numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                    formula_y = re.sub(r'dict', '', numeric_formula)

            rows_to_write.append({
                                    "func_name": func_name,
                                    'formula_d': formula_d,
                                    "formula_m": formula_m,
                                    "formula_y": formula_y,
                                    "roundoff_value":roundoff_value 
                                })
            # print("rows_to_write:", rows_to_write)
        
        return rows_to_write
    except Exception as e:
        return get_exception_response(e)
    

def year_wise_report_print(cnx,machine_id,year,report_for,employee_id,request):

    try:
        print(1)
        groupby = ""
        where = ""
        result = ''
        output = {}
        department_id = ''
        shed_id = ''
        machinetype_id = ''

        def id(machine_id):
            if machine_id !='':
                value = machine_id.split(",")
                if len(value) > 1:
                    if  "all" in value:
                        machine_id = 'all'
                    else:
                        values = tuple(value)
                        machine_id = ",".join(values)
                else:
                    machine_id = value[0]
            return machine_id
        
        machine_id = id(machine_id)

        if machine_id == "" or machine_id == 'all':
            pass
        else:
            
            where += f" and mm.machine_id IN ({machine_id})"

        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]

        if department_id !='' and department_id !=0:
            where += f" and md.department_id ={department_id}"
        if shed_id !=''and shed_id != 0:
            where += f" and ms.shed_id ={shed_id}"
        if machinetype_id !='' and machinetype_id !=0:
            where += f" and mmt.machinetype_id ={machinetype_id}"

        mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        tables_to_union = []
        for month in range(4, 13):
            month_year = f"{mill_month[month]}{year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all()

                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}_12")
                print(month_year)
            else:
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all()

                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
        
        next_year = int(year) + 1
        mill_month = {1: "01", 2: "02", 3: "03"}

        for month in range(1, 4):
            month_year = f"{mill_month[month]}{next_year}"
            print(month_year)
            if report_for == '12to12':
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                result_query = cnx.execute(query).mappings().all() 
                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}_12")
                tables_union_query = " UNION ALL ".join(tables_to_union)
            else:   
                query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                result_query = cnx.execute(query).mappings().all() 
                print("result_query",result_query)
                if len(result_query) > 0:
                    tables_to_union.append(f"select kwh, machine_id,mill_date from ems_v1_completed.dbo.power_{month_year}")
                tables_union_query = " UNION ALL ".join(tables_to_union)
                print("tables_union_query",tables_union_query)

        if len(tables_union_query)==0:
            return _getErrorResponseJson("table not available")
        
        query = text(f'''
            SELECT
                mm.machine_code AS machine_code,
                mm.machine_name AS machine_name,
                ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh,
                FORMAT(min(cp.mill_date), 'MM-yyyy') AS mill_date
            FROM
                ({tables_union_query}) cp
                INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                LEFT JOIN ems_v1.dbo.master_machine_factor mmf ON mmf.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id                   
                LEFT JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id                   
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id                                      
            WHERE
                1=1 {where}
            GROUP BY
                mm.machine_code,
                mm.machine_name,
                MONTH(cp.mill_date), 
                YEAR(cp.mill_date) 
            ORDER BY 
                min(cp.mill_date),  
                min(cp.machine_id)
        ''')

        # print(query)
        rslt = cnx.execute(query).mappings().all()
        if len(rslt)>0:
          output = {}  # Initialize the output dictionary
        
        output_keys = [
            f"04-{year}", f"05-{year}", f"06-{year}",
            f"07-{year}", f"08-{year}", f"09-{year}",
            f"10-{year}", f"11-{year}", f"12-{year}",
            f"01-{next_year}", f"02-{next_year}", f"03-{next_year}"
        ]
        
        for row in rslt:
            machine_code = row['machine_code']
            machine_name = row['machine_name']
            mill_date = row['mill_date']
            kwh = row['kwh']
            
            if machine_code not in output:
                output[machine_code] = {
                    'machine_code': machine_code,
                    'machine_name': machine_name
                }
                for key in output_keys:
                    output[machine_code][key] = 0
            
            output[machine_code][mill_date] = kwh
        
        result = list(output.values())
               
        return result
    except Exception as e:
        return get_exception_response(e)

 
def year_report_print(cnx,year,report_type,request):

    try:
        where = ""
        type = ''
        mill_date = date.today()
        func_name = ''
        formula = ''
        machine_name = ''
        
        kwh = 0
        reverse_kwh = 0
        machine_id = 0
        query = f'''SELECT * FROM [ems_v1].[dbo].[master_energy_calculations] ORDER BY s_no '''
        result = cnx.execute(query).mappings().all()
        if len(result)>0:

            query = text(f'''select mill_date from ems_v1.dbo.master_shifts''')
            shift=cnx.execute(query).mappings().all()
    
            for row in shift :
                mill_date = row['mill_date']  
                
            next_year = int(year) + 1
            mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
            tables_to_union = []
            for month in range(4, 13):
                month_year = f"{mill_month[month]}{year}"

                if report_type == "12to12":
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                    result_query = cnx.execute(query).mappings().all()
                else:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                    result_query = cnx.execute(query).mappings().all()
                if len(result_query) > 0:
                    tables_to_union.append(f"{month_year}")
            
            for month in range(1, 4):
                month_year = f"{mill_month[month]}{next_year}"

                if report_type == "12to12":
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}_12' """
                    result_query = cnx.execute(query).mappings().all() 
                else:
                    query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
                    result_query = cnx.execute(query).mappings().all() 
                if len(result_query) > 0:
                    tables_to_union.append(f"{month_year}")
            print("tables_to_union",tables_to_union)

            if len(tables_to_union)==0:
                 return _getErrorResponseJson("table not available")
            
            if report_type =="12to12":
                type = f"_12 cp"
            else:
                type = f" cp"

            result_dict= {}
            
            for table_name in tables_to_union:
                query = text(f'''
                SELECT
                    mm.machine_name AS machine_name,
                    min(mm.machine_id) machine_id ,
                    ROUND(SUM(case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value  when mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kwh,
                    SUM(case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ) AS reverse_kwh,
                    FORMAT(min(cp.mill_date), 'dd-MM-yyyy') AS mill_date
                FROM
                    ems_v1_completed.dbo.power_{table_name}{type}
                    INNER JOIN ems_v1.dbo.master_machine mm ON mm.machine_id = cp.machine_id
                    LEFT JOIN ems_v1.dbo.master_machine_factor mmf ON mmf.machine_id = mm.machine_id
                WHERE
                    1=1 {where}
                GROUP BY
                    mm.machine_name,
                    DAY(cp.mill_date) 
                ORDER BY 
                    min(cp.mill_date),  
                    min(cp.machine_id)
            ''')
                
                rslt = cnx.execute(query).mappings().all()
                result_dict[table_name] = rslt
            # if len(result_dict[table_name])==0:
            #     return JSONResponse({"iserror": False, "message": "no data available"})
            results = {}
            dict2 = {}
            dict2_r = {}
            roundoff_value_day = 1
            roundoff_value_month = 10
            roundoff_value_year = 1
            formula = 0
            roundoff_values={}

            for table_name, table_data in result_dict.items():
                table_dict = {}
                for row in table_data:
                    machine_id = row['machine_id']
                    mill_date = row['mill_date']
                    kwh = row['kwh']
                    reverse_kwh = row['reverse_kwh']
                
                    if table_name not in dict2:
                        dict2[table_name] = {}
                        dict2_r[table_name] = {}

                    if mill_date not in dict2[table_name]:
                        dict2[table_name][mill_date] = {}
                        dict2_r[table_name][mill_date] = {}

                    if machine_id not in dict2[table_name][mill_date]:
                        dict2[table_name][mill_date][machine_id] = kwh
                        dict2_r[table_name][mill_date][machine_id] = reverse_kwh
                    # print("dict2",dict2)
                    
                    for row in result:
                        func_name = row['function_name']
                        formula = row['formula2']
                        roundoff_value_month = row['roundoff_value']
                        if table_name not in results:
                            results[table_name] = {}

                        if mill_date not in results[table_name]:
                            results[table_name][mill_date] = {}
                        
                        if len(table_data) == 0:
                            formula = 0
                        else:
                            if func_name == 'Power Import':
                                if table_name in dict2 and mill_date in dict2[table_name]:
                                    formula = dict2[table_name][mill_date].get(14, 0)  
                                else:
                                    formula = 0

                            elif func_name == 'Power Export':
                                if table_name in dict2_r and mill_date in dict2_r[table_name]:
                                    formula = dict2_r[table_name][mill_date].get(14, 0)
                                else:
                                    formula = 0
                            else:
                                formula = re.sub(r'dict\[(\d+)\]', lambda match: str(dict2.get(table_name, {}).get(mill_date, {}).get(int(match.group(1)), 0)), formula)
                                formula = eval(formula)
                        results[table_name][mill_date][func_name] = formula
                        roundoff_values[(table_name, mill_date, func_name)] = roundoff_value_month # Add roundoff_value_month

            aggregated_results = {}
        
            for table_name, funcs in results.items():
                for mill_date, formula_result in funcs.items():
                    for func_name, value in formula_result.items():
                        if table_name not in aggregated_results:
                            aggregated_results[table_name] = {}

                        if func_name not in aggregated_results[table_name]:
                            aggregated_results[table_name][func_name] = {"formula": 0.00, "count": 0, "roundoff_value_month": 0}     

                        if aggregated_results[table_name][func_name]["formula"] == 0.00:
                            aggregated_results[table_name][func_name]["formula"] = value
                        else:
                            aggregated_results[table_name][func_name]["formula"] +=value
                        aggregated_results[table_name][func_name]["count"] += 1
                        roundoff_value_month = roundoff_values.get((table_name, mill_date, func_name), 0)
                        aggregated_results[table_name][func_name]["roundoff_value_month"] = roundoff_value_month

            createFolder("YearReport_Log/", "aggregated_results " + f'{aggregated_results}')
            machine_data = {"year_record": {}}

            for table_name, functions in aggregated_results.items():
                for func_name, values in functions.items():
                    kwh = values["formula"]
                    count = values["count"]
                    roundoff_value_month = values["roundoff_value_month"]  
                    
                    if func_name not in machine_data["year_record"]:
                        machine_data["year_record"][func_name] = {}

                    avg_kwh = kwh / count if count > 0 else kwh 
                    machine_data["year_record"][func_name][table_name] = {"roundoff_value_month": roundoff_value_month, "formulas": avg_kwh}
                
            current_month = list(result_dict.keys())[-1]
            machine_data["current_month_record"] = {}

            for row in result_dict[current_month]:
                machine_name = row['machine_name']
                mill_date = row['mill_date']
                kwh = row['kwh']
                reverse_kwh = row['reverse_kwh']

                for row in result:
                    func_name = row['function_name']
                    formula = row['formula2']
                    roundoff_value_day = row['roundoff_value']

                    if current_month not in results:
                        results[current_month] = {}

                    if mill_date not in results[current_month]:
                        results[current_month][mill_date] = {}
                    
                    if roundoff_value_day not in results[current_month][mill_date]:
                        results[current_month][mill_date][roundoff_value_day] = {}
                
                    if len(result_dict[current_month]) == 0:
                        formula_result = 0
                    else:
                        if func_name == 'Power Import':
                            if table_name in dict2 and mill_date in dict2[table_name]:
                                formula = dict2[table_name][mill_date].get(14, 0)
                            else:
                                formula = 0

                        elif func_name == 'Power Export':
                            if table_name in dict2_r and mill_date in dict2_r[table_name]:
                                formula = dict2_r[table_name][mill_date].get(14, 0)
                            else:
                                formula = 0
                        else:
                            formula = re.sub(r'dict\[(\d+)\]', lambda match: str(dict2.get(table_name, {}).get(mill_date, {}).get(int(match.group(1)), 0)), formula)
                            
                    if func_name not in machine_data["current_month_record"]:
                        machine_data["current_month_record"][func_name] = {}
                    
                    machine_data["current_month_record"][func_name][mill_date] = {
                "formula": formula,
                "roundoff_value_day": roundoff_value_day
            }
            para = "case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end "
            union_query = " UNION ALL ".join([f"SELECT cp.machine_id, {para} as data, (case when mmf.kWh = '*' then cp.reverse_kwh * mmf.kWh_value when  mmf.kWh = '/' then cp.reverse_kwh / mmf.kWh_value else cp.reverse_kwh end ) AS reverse_kwh, mm.machine_name FROM ems_v1_completed.dbo.power_{table_name}{type}  left join  [ems_v1].[dbo].[master_machine] mm on mm.machine_id=cp.machine_id left join  [ems_v1].[dbo].[master_machine_factor] mmf on mm.machine_id=mmf.machine_id" for table_name in tables_to_union])
            query4 = f"""
            SELECT 
                pp.machine_id, 
                SUM(pp.data) as total_kwh ,
                SUM(pp.reverse_kwh) AS reverse_kwh,
                min(pp.machine_name) as machine_name
            FROM 
                ({union_query}) AS pp 
                
            GROUP BY 
                pp.machine_id 
            ORDER BY 
                pp.machine_id"""
            createFolder("YearReport_Log/","query for year "+str(query4))
            res_q4 = cnx.execute(query4).mappings().all()
            
            dict_y={}
            dict3 = {}
            dict_r_y = {}
            for row in res_q4:
                dict_y[row['machine_id']] = row['total_kwh']
                dict_r_y[row['machine_id']] = row['reverse_kwh']

            for row in result:
                func_name = row['function_name']
                formula = row['formula2']
                roundoff_value_year = row['roundoff_value']
                if len(res_q4) == 0:
                    formula_y = 0
                else:
                    if func_name == 'Power Import':
                        formula_y = dict_y[14]

                    elif func_name == 'Power Export':
                        formula_y = dict_r_y[14]
                    else:
                        numbers = re.findall(r'\[(\d+)\]', formula)
                        valid_ids = [int(num) for num in numbers if num.isdigit() and int(num) in dict_y]
                        numeric_formula = formula
                        for machine_id in valid_ids:
                            numeric_value = dict_y.get(machine_id, 0)  
                            numeric_formula = numeric_formula.replace(f'[{machine_id}]', str(numeric_value))
                        formula_y = re.sub(r'dict', '', numeric_formula)
                        
                # dict3[func_name] = formula_y
                dict3[func_name] = {
                    'formula_y': formula_y,
                    'roundoff_value_year': roundoff_value_year
                }
            createFolder("YearReport_Log/", "machine_data" + f"{machine_data}")
            response_data = {}
            response_data = {
                "machine_data": machine_data,
                "dict3": dict3,
                "res_q4": res_q4,
                "year": year,
                "next_year": next_year,
                "mill_date": mill_date,
                "report_type":report_type
            }
            # year_wise_excel_report(machine_data, year,next_year,mill_date,dict3,report_type,res_q4)
            # file_path = os.path.join(base_path, f"YearReport-{year}-{next_year}.xlsx")
            # results = f"http://{request.headers['host']}/attachment/YearReport-{year}-{next_year}.xlsx"

        return response_data
    except Exception as e:
        return get_exception_response(e)

def holiday_report(cnx,department_id,shed_id,machinetype_id,machine_id,holiday_type,holiday_year):
    try:

        subquery_union = ''
        tables_to_union = ''
        where = ''
        
        if department_id != '' and department_id != 'all':
            where += f" and md.department_id = {department_id}"

        if shed_id != '' and shed_id != 'all':
            where += f" and ms.shed_id = {shed_id}"

        if machinetype_id != '' and machinetype_id != 'all':
            where += f" and mmt.machinetype_id = {machinetype_id}"

        if machine_id != '' and machine_id != 'all':
            where += f" and mm.machine_id = {machine_id}"
            
        if holiday_type != '' and holiday_type != None:
            where += f" and mhd.holiday_type = '{holiday_type}'"

        mill_month = {1: "01", 2: "02", 3: "03", 4: "04", 5: "05", 6: "06",7: "07", 8: "08", 9: "09", 10: "10", 11: "11", 12: "12"}
        tables_to_union = []

        for month in range(1, 13):
            month_year = f"{mill_month[month]}{holiday_year}"
            print(month_year)
            query = f"""SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_CATALOG = 'ems_v1_completed' AND TABLE_SCHEMA = 'dbo' AND TABLE_NAME = 'power_{month_year}' """
            result_query = cnx.execute(query).mappings().all()

            if len(result_query) > 0:
                tables_to_union.append(f"select mill_date, mill_shift, machine_kWh, master_kwh, kWh, machine_id from ems_v1_completed.dbo.power_{month_year}")

            if len(tables_to_union) == 0:
                return _getErrorResponseJson("power table not available...")    
            subquery_union = " UNION ALL ".join(tables_to_union) 

        query = text(f'''
                select 
                    cp.mill_date,
                    cp.mill_shift,
                    FORMAT(CONVERT(date, cp.mill_date), 'MMMM') AS month,
                    md.department_code,
                    md.department_name,
                    ms.shed_code,
                    ms.shed_name,
                    mmt.machinetype_code,
                    mmt.machinetype_name,
                    mm.machine_code,
                    mm.machine_name,
                    cp.machine_id,
                    ROUND((case when mmf.machine_kWh = '*' then cp.machine_kWh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.machine_kWh / mmf.machine_kWh_value else cp.machine_kWh end) ,2) AS end_kwh,
                    ROUND((case when mmf.machine_kWh = '*' then cp.master_kwh * mmf.machine_kWh_value when  mmf.machine_kWh = '/' then cp.master_kwh / mmf.machine_kWh_value else cp.master_kwh end ),2) AS start_kwh,
                    ROUND((case when mmf.kWh = '*' then cp.kWh * mmf.kWh_value when  mmf.kWh = '/' then cp.kWh / mmf.kWh_value else cp.kWh end ),2) AS kWh,
                    mhd.description,
                    mhd.holiday_type
                    from
                        ({subquery_union}) as cp
                    INNER JOIN [ems_v1].[dbo].[master_machine] mm ON mm.machine_id = cp.machine_id
                    INNER JOIN [ems_v1].[dbo].[master_department] md ON md.department_id = mm.department_id
                    INNER JOIN [ems_v1].[dbo].[master_shed] ms ON ms.shed_id = mm.shed_id
                    INNER JOIN [ems_v1].[dbo].[master_machinetype] mmt ON mmt.machinetype_id = mm.machinetype_id 
                    INNER JOIN [ems_v1].[dbo].[master_holiday_machine] mhm ON mhm.machine_id = mm.machine_id 
                    INNER JOIN [ems_v1].[dbo].[master_holiday_date] mhd ON mhd.holiday_date = cp.mill_date 
                    INNER JOIN [ems_v1].[dbo].[master_holiday] mh ON mh.id = mhm.ref_id and mh.id = mhd.ref_id and mh.status = 'active'
                    LEFT JOIN [ems_v1].[dbo].[master_machine_factor] mmf ON mm.machine_id = mmf.machine_id     
                    where 1=1 and mh.holiday_year = {holiday_year} {where}
                    order by cp.mill_date, cp.mill_shift, mm.machine_order
                    ''')
        createFolder("holiday/","query forholiday"+str(query))
        data = cnx.execute(query).mappings().all()
                
        return data
    except Exception as e:
        return get_exception_response(e)
    
def alarmreport(cnx,company_id,department_id,shed_id,machinetype_id,machine_id,report_for,period_id,from_date,to_date,shift_id, employee_id  ):

    try:

        if  employee_id != '':
            query = text(f'''select * from ems_v1.dbo.master_employee where employee_id = {employee_id}''')
            res = cnx.execute(query).mappings().all()
            if len(res)>0:
                for row in res:
                    department_id = row["department_id"]
                    shed_id = row["shed_id"]
                    machinetype_id = row["machinetype_id"]
        start_time = date.today()
       
        where = ""         
        query = text(f'''SELECT * FROM ems_v1.dbo.master_shifts WHERE status = 'active' ''')
        data1 = cnx.execute(query).fetchall()
        mill_date = date.today()
        mill_shift = 0        
    
        if len(data1) > 0:
           for shift_record in data1:
              mill_date = shift_record["mill_date"]
              mill_shift = shift_record["mill_shift"]            
              print(mill_date)
        if period_id == "cur_shift":          

            where += f'''  pa.mill_date = '{mill_date}' AND pa.mill_shift = '{mill_shift}' ''' 
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) '''
            
        elif period_id == "sel_shift":            
            if from_date is None:
                return _getErrorResponseJson("from date is required")
            if shift_id is None:
                return _getErrorResponseJson("shift_id is required")                            
            where += f'''  pa.mill_date = '{parse_date(from_date)}' AND pa.mill_shift = '{shift_id}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) ''' 

        elif period_id == "sel_date":            
            if from_date is None:
                 return _getErrorResponseJson("from date is required")                     
            where += f'''pa.mill_date = '{parse_date(from_date)}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time)'''
            
        elif period_id == "from_to":            
            if from_date is None:
                return _getErrorResponseJson("from date is required")
            if to_date is None:
                return _getErrorResponseJson("to_date is required")            
                    
            where += f''' pa.mill_date  >= '{parse_date(from_date)}' and pa.mill_date <= '{parse_date(to_date)}' '''
            duration = f'''DATEDIFF(second, pa.start_time, pa.stop_time) '''
            if shift_id is not None:                
                where += f''' and pa.mill_shift = '{shift_id}' ''' 
                       
        elif period_id == "live_alarm":
            where += f''' pa.start_time <> '1900-01-01 00:00:00'  and pa.stop_time is Null '''  
            sql = text(f'''SELECT TOP 1 FORMAT(start_time, 'yyyy-MM-dd HH:mm:ss') as start_time FROM ems_v1.dbo.present_alarm ORDER BY start_time DESC''')

            data = cnx.execute(sql).fetchall()
            duration = f'''DATEDIFF(second, pa.start_time, getdate())'''
            for i in data:
                start_time = i['start_time']
            print(start_time)

            sql1= text(f''' UPDATE ems_v1.dbo.master_company
            SET alarm_status = CASE WHEN alarm_last_time = '{start_time}' THEN alarm_status ELSE 1 END,
                alarm_last_time = '{start_time}'
            WHERE company_id = '{company_id}'
        ''')

            cnx.execute(sql1)
            cnx.commit()
            query2=text(f'''select * from ems_v1.dbo.master_company where company_id = '{company_id}' and alarm_status = 1''')
            data1 = cnx.execute(query2).fetchall() 
        else:
             return _getErrorResponseJson("invalid period id")   
    
        if machine_id is not None and machine_id != 'all':
            where += f" and mm.machine_id = '{machine_id}' "   

        if company_id !='' and company_id !="0":
            where += f" and mc.company_id = '{company_id}' "

        if department_id !='' and department_id !="0":
            where += f" and md.department_id = '{department_id}' " 

        if shed_id !='' and  shed_id !="0":
            where += f" and ms.shed_id = '{shed_id}' "   

        if machinetype_id !='' and machinetype_id!="0":
            where += f" and mmt.machinetype_id = '{machinetype_id}' "  

        groupby = ""  
        if report_for == "summary":
            sql = text(f'''
                SELECT 
                    min(ma.alarm_name) as alarm_name,		    
                    min(pa.parameter_name) as parameter_name,
                    min(mm.machine_name) as machine_name,
                    sum({duration}) as duration
                ''')
            groupby = f'''group by ma.alarm_name '''
            
        else:
            sql = text(f'''
                SELECT 
                    mm.machine_code,
                    mm.machine_name,
                    ma.alarm_name,
                    ma.parameter_name,
                    pa.start_time,
                    pa.stop_time,
                    {duration} as duration,
                    pa.description
                ''')
        query = text(f''' 
            {sql}
            FROM 
                ems_v1.dbo.master_alarm_target ma 
                INNER JOIN ems_v1.dbo.present_alarm pa on pa.alarm_target_id = ma.alarm_target_id
                INNER JOIN ems_v1.dbo.master_machine mm on pa.machine_id = mm.machine_id
                LEFT JOIN [ems_v1].[dbo].[master_company] mc on mc.company_id=ma.company_id
                LEFT JOIN [ems_v1].[dbo].[master_branch] mb on mb.branch_id=ma.branch_id
                LEFT JOIN [ems_v1].[dbo].[master_department] md on md.department_id=ma.department_id
                LEFT JOIN [ems_v1].[dbo].[master_shed] ms on ms.shed_id=ma.shed_id
                LEFT JOIN [ems_v1].[dbo].[master_machinetype] mmt on mmt.machinetype_id=ma.machinetype_id
            WHERE  {where} {groupby}
         ''')

        data = cnx.execute(query).fetchall() 
                            
        return data
    except Exception as e:
        return get_exception_response(e)
    
def hour_wise_analysis_report(cnx,machine_id,from_time,to_time):

    try:

        mill_date = date.today()
        mill_shift = 0
        where = ''
        
        query=text(f'''SELECT * FROM [ems_v1].[dbo].[master_shifts] WHERE status='active' ''')
        data1 = cnx.execute(query).mappings().all()
        mill_date = date.today()
        mill_shift = 0  

        if len(data1) > 0:
            for shift_record in data1:
                mill_date = shift_record["mill_date"]
                mill_shift = shift_record["mill_shift"]  
                    
        where += f" and cp.mill_date = '{mill_date}' and cp.mill_shift ='{mill_shift}' "
    
        if from_time !='':
            where += f" and FORMAT(cp.created_on ,'HH:mm:ss')>='{from_time}' "
    
        if to_time !='':
            where += f" and FORMAT(cp.created_on ,'HH:mm:ss')<='{to_time}' "
            
        query=text(f'''
            SELECT
                mm.machine_id,
                min(mm.machine_code) as machine_code,
                min(mm.machine_name) as machine_name,
                MIN(cp.created_on) AS start_time,
                MAX(cp.created_on) AS end_time,
                min(cp.machine_kwh) as start_kwh,
                max(cp.machine_kwh) as end_kwh,
                max(cp.machine_kwh) - min(cp.machine_kwh) as kwh
            FROM
                ems_v1.dbo.current_power_analysis cp
                INNER JOIN [ems_v1].[dbo].[master_machine] mm ON mm.machine_id = cp.machine_id
            WHERE mm.machine_id in ({machine_id}) {where}
            group by mm.machine_id
          ''')  
        data = cnx.execute(query).fetchall() 
                            
        return data
    except Exception as e:
        return get_exception_response(e)
    